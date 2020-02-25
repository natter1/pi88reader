"""
todo: check pandas
"""
from openpyxl import Workbook
from openpyxl.styles import Font

from pi88reader.pi88_importer import PI88Measurement, SegmentType


def main():
    filename = '..\\resources\\quasi_static_12000uN.tdm'
    filename = '..\\resources\\AuSn_Creep\\1000uN 01 LC.tdm'
    measurement = PI88Measurement(filename)

    to_excel = PI88ToExcel(measurement)
    to_excel.write("delme.xlsx")


class PI88ToExcel:
    def __init__(self, pi88_measurement):
        self.measurement = pi88_measurement
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def write(self, filename):
        self.add_sheet_quasi_static_data()  # self.workbook.active)
        self.add_sheet_segment_data()
        self.workbook.save(filename=filename)

    def add_sheet_quasi_static_data(self):
        wb = self.workbook
        #mws_title = self.measurement.filename.split('.')[2].split('\\')[2]
        ws_title = self.measurement.filename.split('\\')[-1].split('.')[0]
        ws = wb.create_sheet(title=ws_title)
        data = self.measurement.get_quasi_static_curve()
        self.write_data(ws, data)

    def add_sheet_segment_data(self):
        wb = self.workbook
        ws_title = "segments"
        ws = wb.create_sheet(title=ws_title)

        ws.cell(row=1, column=1).value = "LOAD:"
        data = self.measurement.get_segment_curve(SegmentType.LOAD)
        self.write_data(ws, data, row=1, col=2)

        ws.cell(row=1, column=5).value = "HOLD:"
        data = self.measurement.get_segment_curve(SegmentType.HOLD)
        self.write_data(ws, data, row=1, col=6)

        ws.cell(row=1, column=9).value = "UNLOAD:"
        data = self.measurement.get_segment_curve(SegmentType.UNLOAD)
        self.write_data(ws, data, row=1, col=10)

    @staticmethod
    def write_row(ws, data, row, col):
        font = Font(bold=True)
        for i, value in enumerate(data):
            ws.cell(row=row, column=col+i).value = value
            ws.cell(row=row, column=col + i).font = font

    @staticmethod
    def write_cols(ws, data, row, col):
        for i, value in enumerate(data[0]):
            for j, column in enumerate(data):
                ws.cell(row=row+i, column=col+j).value = column[i]

    def write_data(self, ws, data, row=1, col=1):
        header = data[0]
        if header:
            self.write_row(ws, header, row, col)
            row += 1
        self.write_cols(ws, data[1:], row, col)
        # for i, value in enumerate(data[1]):
        #     for j, column in enumerate(data[1:]):
        #         ws.cell(row=row+i, column=col+j).value = column[i]


if __name__ == "__main__":
    main()
